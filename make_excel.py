import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Packing List"

# Header Information
ws.merge_cells("A1:J1")
ws["A1"] = "PACKING LIST: PL- 25/ASN/TM/VII/26 (CONTAINER: BEAU 5231653/ID48136AA)"
ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="1B5E20")
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 30

# Column Headers
headers = [
    "ลำดับ",
    "No. Item (Part Number)",
    "No. BDL (มัดที่)",
    "ขนาด / Description (MM)",
    "ยอดตาม PL (PCS)",
    "ปริมาตร (M3)",
    "ยอดรับเข้าแล้ว (PCS)",
    "ยอดคงเหลือ (PCS)",
    "สถานะการตรวจรับ",
    "อัปเดตล่าสุด"
]

header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
thin_border = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0')
)

ws.append([]) # Row 2 empty
ws.append(headers) # Row 3
ws.row_dimensions[3].height = 25

for col_idx in range(1, len(headers) + 1):
    cell = ws.cell(row=3, column=col_idx)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Master Data from PDF
data = [
  ("8163986", "1", "16 x 254 x 524 MM", 250, 0.532),
  ("8156594", "-", "16 x 162 x 644 MM", 250, 0.417),
  ("8156623", "2", "16 x 162 x 872 MM", 438, 0.990),
  ("8156623", "3", "16 x 162 x 872 MM", 232, 0.524),
  ("8156614", "4", "16 x 162 x 796 MM", 438, 0.904),
  ("8156614", "5", "16 x 162 x 796 MM", 312, 0.644),
  ("8156604", "6", "16 x 162 x 720 MM", 438, 0.817),
  ("8156604", "7", "16 x 162 x 720 MM", 312, 0.582),
  ("8156585", "8", "16 x 162 x 568 MM", 750, 1.104),
  ("8156521", "9", "16 x 162 x 524 MM", 876, 1.190),
  ("8156521", "10", "16 x 162 x 524 MM", 876, 1.190),
  ("8156521", "11", "16 x 162 x 524 MM", 876, 1.190),
  ("8156521", "12", "16 x 162 x 524 MM", 876, 1.190),
  ("8156521", "13", "16 x 162 x 524 MM", 856, 1.163),
  ("8156578", "14", "16 x 162 x 491 MM", 320, 0.407),
  ("8156374", "-", "16 x 162 x 474 MM", 404, 0.496),
  ("8156374", "15", "16 x 162 x 474 MM", 876, 1.076),
  ("8156666", "16", "16 x 86 x 872 MM", 320, 0.384),
  ("8156663", "-", "16 x 86 x 796 MM", 480, 0.526),
  ("8156544", "17", "16 x 86 x 822 MM", 320, 0.362),
  ("8156535", "-", "16 x 86 x 746 MM", 480, 0.493),
  ("8156659", "18", "16 x 86 x 720 MM", 960, 0.951),
  ("8156646", "19", "16 x 86 x 568 MM", 520, 0.406),
  ("8156643", "-", "16 x 86 x 491 MM", 640, 0.432),
  ("8156575", "20", "16 x 86 x 524 MM", 1608, 1.159),
  ("8156575", "21", "16 x 86 x 524 MM", 1608, 1.159),
  ("8156575", "22", "16 x 86 x 524 MM", 1264, 0.911),
  ("8156565", "23", "16 x 86 x 474 MM", 1608, 1.049),
  ("8156565", "24", "16 x 86 x 474 MM", 1592, 1.038),
  ("6601644", "25", "15 x 279.4 x 520.7 MM", 528, 1.152),
  ("6601644", "26", "15 x 279.4 x 520.7 MM", 528, 1.152),
  ("6601644", "27", "15 x 279.4 x 520.7 MM", 144, 0.314),
  ("6601640", "-", "15 x 228.6 x 673.1 MM", 520, 1.200),
  ("6601637", "28", "15 x 228.6 x 520.7 MM", 664, 1.186),
  ("6601637", "29", "15 x 228.6 x 520.7 MM", 664, 1.186),
  ("6601637", "30", "15 x 228.6 x 520.7 MM", 242, 0.432),
  ("6601636", "-", "15 x 228.6 x 444.5 MM", 386, 0.588),
  ("6601636", "31", "15 x 228.6 x 444.5 MM", 664, 1.012),
  ("8843617", "32", "15 x 228.6 x 431.8 MM", 295, 0.437),
  ("6601608", "-", "15 x 101.6 x 520.7 MM", 1100, 0.873),
  ("6601628", "33", "15 x 177.8 x 673.1 MM", 315, 0.565),
  ("6601612", "-", "15 x 101.6 x 647.7 MM", 570, 0.563),
  ("6601609", "-", "15 x 101.6 x 566.7 MM", 570, 0.492),
  ("6601625", "34", "15 x 177.8 x 520.7 MM", 864, 1.200),
  ("6601625", "35", "15 x 177.8 x 520.7 MM", 816, 1.133),
  ("6601608", "36", "15 x 101.6 x 520.7 MM", 1500, 1.190),
  ("6601608", "37", "15 x 101.6 x 520.7 MM", 1500, 1.190),
  ("6601608", "38", "15 x 101.6 x 520.7 MM", 1500, 1.190)
]

for idx, (item, bdl, dim, qty, vol) in enumerate(data, start=1):
    row_num = idx + 3
    row_val = [
        idx,
        item,
        bdl,
        dim,
        qty,
        vol,
        0,
        f"=E{row_num}-G{row_num}",
        f'=IF(G{row_num}=0,"⏳ ยังไม่รับ",IF(G{row_num}>=E{row_num},"✅ ครบถ้วน","🟡 รับบางส่วน"))',
        "-"
    ]
    ws.append(row_val)
    
    # Styling row
    for c_idx in range(1, len(row_val) + 1):
        cell = ws.cell(row=row_num, column=c_idx)
        cell.border = thin_border
        if c_idx in [1, 2, 3, 9, 10]:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif c_idx in [5, 6, 7, 8]:
            cell.alignment = Alignment(horizontal="right", vertical="center")

# Total Row
tot_row = len(data) + 4
ws.cell(row=tot_row, column=1, value="รวมทั้งหมด (Total)").font = Font(name="Calibri", size=11, bold=True)
ws.cell(row=tot_row, column=5, value=f"=SUM(E4:E{tot_row-1})").font = Font(name="Calibri", size=11, bold=True)
ws.cell(row=tot_row, column=6, value=f"=SUM(F4:F{tot_row-1})").font = Font(name="Calibri", size=11, bold=True)
ws.cell(row=tot_row, column=7, value=f"=SUM(G4:G{tot_row-1})").font = Font(name="Calibri", size=11, bold=True)
ws.cell(row=tot_row, column=8, value=f"=SUM(H4:H{tot_row-1})").font = Font(name="Calibri", size=11, bold=True)

tot_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
for c in range(1, len(headers) + 1):
    cell = ws.cell(row=tot_row, column=c)
    cell.fill = tot_fill
    cell.border = thin_border

# Auto adjust column widths
for col in ws.columns:
    max_len = max(len(str(cell.value or '')) for cell in col)
    col_letter = openpyxl.utils.get_column_letter(col[0].column)
    ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

out_file = r"c:\Users\Asus 444\Desktop\P&T สิงหนคร (ตัวแทนจำหน่ายรถมือสอง )\whatsapp-bot\Data Invoice\Packing_List_BEAU5231653.xlsx"
wb.save(out_file)
print(f"Saved Excel to {out_file}")
